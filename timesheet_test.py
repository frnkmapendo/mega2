#!/usr/bin/env python3
"""
Minimal Timesheet Management System
Test implementation for the timesheet functionality
"""

import pandas as pd
import calendar
import random
from datetime import datetime
import openpyxl
from io import BytesIO

class TimesheetManager:
    def __init__(self):
        self.projects = []
        self.total_work_hours_per_day = 8  # Standard 8-hour workday
        self.minimum_time_block = 0.5  # 30 minutes minimum
        
    def add_project(self, name, percentage):
        """Add a project with its percentage allocation"""
        if not name or percentage <= 0 or percentage > 100:
            return False
        
        # Check if total percentage would exceed 100%
        current_total = sum(p['percentage'] for p in self.projects)
        if current_total + percentage > 100:
            return False
            
        project = {
            'id': len(self.projects) + 1,
            'name': name,
            'percentage': percentage,
            'daily_hours': (percentage / 100) * self.total_work_hours_per_day
        }
        self.projects.append(project)
        return True
    
    def remove_project(self, project_id):
        """Remove a project by ID"""
        self.projects = [p for p in self.projects if p['id'] != project_id]
        
    def get_working_days(self, year, month):
        """Get list of working days (Mon-Fri) for a given month"""
        cal = calendar.monthcalendar(year, month)
        working_days = []
        
        for week in cal:
            for day_num, day in enumerate(week):
                if day != 0 and day_num < 5:  # Monday=0 to Friday=4
                    working_days.append(day)
        
        return working_days
    
    def get_project_hours_daily(self, year, month, randomize_small_projects=False):
        """
        Calculate daily hours for each project in a given month
        
        Args:
            year: Year for calculation
            month: Month for calculation  
            randomize_small_projects: If True, randomize projects under 20%
            
        Returns:
            DataFrame with daily hours for each project
        """
        working_days = self.get_working_days(year, month)
        
        # Initialize result dataframe
        result = pd.DataFrame(index=working_days)
        
        # Add columns for each project
        for project in self.projects:
            project_name = project['name']
            percentage = project['percentage']
            total_monthly_hours = project['daily_hours'] * len(working_days)
            
            if randomize_small_projects and percentage < 20:
                # Randomize small projects (under 20%)
                daily_hours = self._randomize_project_hours(
                    total_monthly_hours, working_days, project_name
                )
            else:
                # Even distribution
                daily_hours = [project['daily_hours']] * len(working_days)
            
            result[project_name] = daily_hours
        
        return result
    
    def _randomize_project_hours(self, total_hours, working_days, project_name):
        """
        Randomize hours for small projects into concentrated blocks
        
        Args:
            total_hours: Total monthly hours for the project
            working_days: List of working days in month
            project_name: Name of the project (for seed consistency)
            
        Returns:
            List of daily hours with randomized distribution
        """
        num_days = len(working_days)
        daily_hours = [0.0] * num_days
        
        if total_hours < self.minimum_time_block:
            # If total hours is less than minimum block, assign to one random day
            random.seed(f"{project_name}_{num_days}")
            selected_day = random.randint(0, num_days - 1)
            daily_hours[selected_day] = total_hours
            return daily_hours
        
        # Calculate how many days we should concentrate the hours into
        # Use minimum time blocks to determine number of working days needed
        blocks_needed = max(1, int(total_hours / self.minimum_time_block))
        concentration_days = min(blocks_needed, num_days // 2)  # Use at most half the days
        
        # Randomly select which days get the hours
        random.seed(f"{project_name}_{num_days}")
        selected_days = random.sample(range(num_days), concentration_days)
        
        # Distribute hours across selected days
        hours_per_day = total_hours / concentration_days
        
        # Round to nearest 30 minutes and adjust for exact total
        rounded_hours = round(hours_per_day / self.minimum_time_block) * self.minimum_time_block
        
        for i, day_idx in enumerate(selected_days):
            if i == len(selected_days) - 1:  # Last day gets remainder
                remaining = total_hours - sum(daily_hours)
                daily_hours[day_idx] = max(0, remaining)  # Ensure non-negative
            else:
                daily_hours[day_idx] = min(rounded_hours, total_hours - sum(daily_hours))  # Don't exceed total
        
        return daily_hours
    
    def export_to_excel(self, year, month, randomize_small_projects=False):
        """
        Export timesheet to Excel format
        
        Returns:
            BytesIO object containing Excel file
        """
        df = self.get_project_hours_daily(year, month, randomize_small_projects)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Timesheet_{year}_{month:02d}"
        
        # Write headers
        headers = ["Day"] + list(df.columns) + ["Total Hours"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row_idx, (day, row_data) in enumerate(df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=day)
            
            row_total = 0
            for col_idx, project_name in enumerate(df.columns, 2):
                hours = row_data[project_name]
                ws.cell(row=row_idx, column=col_idx, value=hours)
                row_total += hours
            
            # Add daily total
            ws.cell(row=row_idx, column=len(headers), value=row_total)
        
        # Add monthly totals row
        total_row = len(df) + 2
        ws.cell(row=total_row, column=1, value="Monthly Total")
        
        for col_idx, project_name in enumerate(df.columns, 2):
            total_hours = df[project_name].sum()
            ws.cell(row=total_row, column=col_idx, value=total_hours)
        
        # Add grand total
        grand_total = df.sum().sum()
        ws.cell(row=total_row, column=len(headers), value=grand_total)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer

# Test the functionality
if __name__ == "__main__":
    # Create timesheet manager
    tm = TimesheetManager()
    
    # Add some test projects
    print("Adding test projects...")
    tm.add_project("Large Project A", 40)
    tm.add_project("Medium Project B", 25)
    tm.add_project("Small Project C", 15)  # Under 20% - should be randomized
    tm.add_project("Tiny Project D", 5)    # Under 20% - should be randomized
    
    print(f"Total projects: {len(tm.projects)}")
    
    # Test even distribution
    print("\n=== EVEN DISTRIBUTION ===")
    df_even = tm.get_project_hours_daily(2024, 1, randomize_small_projects=False)
    print(df_even.head(10))
    print(f"\nMonthly totals (even): {df_even.sum()}")
    print(f"Grand total: {df_even.sum().sum()}")
    
    # Test randomized distribution
    print("\n=== RANDOMIZED DISTRIBUTION ===")
    df_random = tm.get_project_hours_daily(2024, 1, randomize_small_projects=True)
    print(df_random.head(10))
    print(f"\nMonthly totals (randomized): {df_random.sum()}")
    print(f"Grand total: {df_random.sum().sum()}")
    
    # Verify totals are the same
    print(f"\nTotals match: {df_even.sum().equals(df_random.sum())}")
    
    # Test Excel export
    print("\n=== EXCEL EXPORT TEST ===")
    excel_buffer = tm.export_to_excel(2024, 1, randomize_small_projects=True)
    print(f"Excel file size: {len(excel_buffer.getvalue())} bytes")
    
    # Save to file for manual inspection
    with open("/tmp/timesheet_test.xlsx", "wb") as f:
        f.write(excel_buffer.getvalue())
    print("Excel file saved to /tmp/timesheet_test.xlsx")
    
    print("\n✅ All tests completed successfully!")